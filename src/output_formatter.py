"""
Unified Output Formatter for Compliance Results

This module provides a consistent API for generating formatted compliance reports
in multiple output formats (JSON, CSV, HTML, etc.) from both individual document 
validation results and compliance matrices across documents.

The formatter supports both static and dynamic evaluation modes and ensures
all outputs follow consistent schemas with essential compliance information.
"""

import json
import csv
import os
import io
import base64
import tempfile
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Tuple
from enum import Enum
from datetime import datetime
import logging

from .output_format import (
    ValidationStatus, 
    ValidationResult,
    ValidationResultFormatter
)

from .compliance_evaluator import (
    ComplianceLevel,
    ComplianceResult,
    DocumentComplianceReport
)

from .compliance_matrix_generator import (
    ComplianceMatrixGenerator,
    OutputFormat as MatrixOutputFormat,
    VisualizationStyle
)


class OutputFormat(str, Enum):
    """Supported output formats for compliance reports"""
    JSON = "json"
    CSV = "csv"
    HTML = "html"
    MARKDOWN = "markdown"
    EXCEL = "excel"


class OutputType(str, Enum):
    """Types of compliance output"""
    DOCUMENT = "document"      # Single document output
    MATRIX = "matrix"          # Matrix comparing multiple documents/requirements
    SUMMARY = "summary"        # Summary statistics only


class OutputFormatter:
    """
    Unified formatter for compliance results with support for multiple output formats.
    
    This class provides a consistent API for generating formatted compliance reports
    from both individual document validation results and compliance matrices.
    """
    
    def __init__(
        self,
        include_details: bool = True,
        include_justifications: bool = True,
        include_confidence: bool = True,
        include_metadata: bool = True,
        visualization_style: str = "color",
        logger: Optional[logging.Logger] = None
    ):
        """
        Initialize the output formatter.
        
        Args:
            include_details: Whether to include detailed information in outputs
            include_justifications: Whether to include justification text
            include_confidence: Whether to include confidence scores
            include_metadata: Whether to include metadata
            visualization_style: Visual style for HTML/Excel outputs
                                (text, symbol, color, colorblind)
            logger: Optional logger for recording formatter operations
        """
        self.include_details = include_details
        self.include_justifications = include_justifications
        self.include_confidence = include_confidence
        self.include_metadata = include_metadata
        self.visualization_style = self._parse_visualization_style(visualization_style)
        self.logger = logger or logging.getLogger(__name__)
        
        # Initialize matrix generator with our settings
        self.matrix_generator = ComplianceMatrixGenerator(
            evaluator=None,  # Will be supplied when needed
            visualization_style=self.visualization_style,
            include_justifications=include_justifications,
            include_confidence=include_confidence,
            include_metadata=include_metadata
        )
    
    def _parse_visualization_style(self, style: str) -> VisualizationStyle:
        """Convert string style to VisualizationStyle enum"""
        try:
            return VisualizationStyle(style)
        except ValueError:
            # Default to color if invalid style provided
            self.logger.warning(f"Invalid visualization style '{style}', defaulting to 'color'")
            return VisualizationStyle.COLOR
    
    def format_document_result(
        self,
        result: Union[ValidationResult, ComplianceResult, Dict[str, Any]],
        output_format: OutputFormat = OutputFormat.JSON,
        output_path: Optional[Path] = None
    ) -> Union[Dict[str, Any], str, Path]:
        """
        Format a single document compliance result.
        
        Args:
            result: The validation or compliance result to format
            output_format: Desired output format
            output_path: Optional path to save the output
            
        Returns:
            Based on output_format:
              - JSON: Dictionary containing the formatted result
              - CSV/HTML/MARKDOWN: String containing the formatted output
              - EXCEL: Path to the saved Excel file
        """
        if isinstance(result, dict):
            # If provided as a dictionary, assume it's already a properly structured result
            formatted_result = result
        elif isinstance(result, ValidationResult):
            # Convert ValidationResult to dictionary using existing formatter
            formatted_result = ValidationResultFormatter.to_dict(result)
        elif isinstance(result, ComplianceResult):
            # Convert ComplianceResult to our standard format
            formatted_result = self._convert_compliance_result(result)
        else:
            raise TypeError(f"Unsupported result type: {type(result)}")
        
        # Format based on the requested output format
        if output_format == OutputFormat.JSON:
            if output_path:
                self._save_json(formatted_result, output_path)
            return formatted_result
            
        elif output_format == OutputFormat.CSV:
            csv_content = self._convert_document_to_csv(formatted_result)
            if output_path:
                self._save_text(csv_content, output_path)
            return csv_content
            
        elif output_format == OutputFormat.HTML:
            html_content = self._convert_document_to_html(formatted_result)
            if output_path:
                self._save_text(html_content, output_path)
            return html_content
            
        elif output_format == OutputFormat.MARKDOWN:
            md_content = self._convert_document_to_markdown(formatted_result)
            if output_path:
                self._save_text(md_content, output_path)
            return md_content
            
        elif output_format == OutputFormat.EXCEL:
            if not output_path:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
                    output_path = Path(temp.name)
            
            self._convert_document_to_excel(formatted_result, output_path)
            return output_path
            
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    
    def format_compliance_matrix(
        self,
        reports: Dict[str, Union[DocumentComplianceReport, ValidationResult, Dict]],
        output_format: OutputFormat = OutputFormat.JSON,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_order: str = "ascending",
        output_path: Optional[Path] = None
    ) -> Union[Dict[str, Any], str, Path]:
        """
        Format a compliance matrix from multiple document reports.
        
        Args:
            reports: Dictionary mapping document IDs to compliance reports
            output_format: Desired output format
            filters: Optional filters to apply
            sort_by: Optional field to sort by
            sort_order: Sorting order (ascending/descending)
            output_path: Optional path to save the output
            
        Returns:
            Based on output_format:
              - JSON: Dictionary containing the matrix
              - CSV/HTML/MARKDOWN: String containing the formatted output
              - EXCEL: Path to the saved Excel file
        """
        # Convert reports to common format if needed
        normalized_reports = {}
        for doc_id, report in reports.items():
            if isinstance(report, dict):
                # Assume it's already a DocumentComplianceReport dict
                normalized_reports[doc_id] = report
            elif isinstance(report, ValidationResult):
                # Convert ValidationResult to DocumentComplianceReport
                normalized_reports[doc_id] = self._convert_validation_result(report)
            elif isinstance(report, DocumentComplianceReport):
                # Already the right type
                normalized_reports[doc_id] = report
            else:
                raise TypeError(f"Unsupported report type for document {doc_id}: {type(report)}")
        
        # Map our OutputFormat to the matrix generator's format
        matrix_format = self._map_output_format(output_format)
        
        # Use the compliance matrix generator to create the matrix
        result = self.matrix_generator.generate_matrix(
            reports=normalized_reports,
            output_format=matrix_format,
            filters=filters,
            sort_by=sort_by,
            sort_order=sort_order,
            output_path=output_path
        )
        
        return result
    
    def generate_report(
        self,
        data: Union[Dict[str, Any], List[Dict[str, Any]]],
        output_type: OutputType,
        output_format: OutputFormat = OutputFormat.JSON,
        output_path: Optional[Path] = None,
        **kwargs
    ) -> Union[Dict[str, Any], str, Path]:
        """
        Generate a compliance report from data.
        
        This is the main entry point for generating reports, handling
        both single document results and matrices.
        
        Args:
            data: The data to format (document result or dictionary of results)
            output_type: Type of output (document, matrix, or summary)
            output_format: Desired output format
            output_path: Optional path to save the output
            **kwargs: Additional options for specific output types
            
        Returns:
            The formatted output in the requested format
        """
        if output_type == OutputType.DOCUMENT:
            # Handle single document output
            if not isinstance(data, (dict, ValidationResult, ComplianceResult)):
                raise TypeError("Data for DOCUMENT output type must be a dictionary, "
                               "ValidationResult, or ComplianceResult")
            
            return self.format_document_result(
                result=data,
                output_format=output_format,
                output_path=output_path
            )
            
        elif output_type == OutputType.MATRIX:
            # Handle matrix output
            if not isinstance(data, dict):
                raise TypeError("Data for MATRIX output type must be a dictionary "
                               "mapping document IDs to results")
            
            return self.format_compliance_matrix(
                reports=data,
                output_format=output_format,
                filters=kwargs.get('filters'),
                sort_by=kwargs.get('sort_by'),
                sort_order=kwargs.get('sort_order', 'ascending'),
                output_path=output_path
            )
            
        elif output_type == OutputType.SUMMARY:
            # Handle summary output
            if isinstance(data, dict) and not any(isinstance(v, dict) for v in data.values()):
                # Single document summary
                return self._generate_document_summary(
                    data, output_format, output_path
                )
            else:
                # Multiple document summary
                return self._generate_matrix_summary(
                    data, output_format, output_path
                )
                
        else:
            raise ValueError(f"Unsupported output type: {output_type}")
    
    def _convert_compliance_result(self, result: ComplianceResult) -> Dict[str, Any]:
        """Convert a ComplianceResult to our standard document output format"""
        # Map ComplianceLevel to ValidationStatus
        status_map = {
            ComplianceLevel.FULLY_COMPLIANT: ValidationStatus.PASSED,
            ComplianceLevel.PARTIALLY_COMPLIANT: ValidationStatus.PARTIAL,
            ComplianceLevel.NON_COMPLIANT: ValidationStatus.FAILED,
            ComplianceLevel.NOT_APPLICABLE: ValidationStatus.UNKNOWN,
            ComplianceLevel.INDETERMINATE: ValidationStatus.UNKNOWN
        }
        
        # Extract document info
        doc_info = result.details.get('document_info', {})
        doc_id = doc_info.get('id', 'unknown')
        doc_name = doc_info.get('name', doc_id)
        doc_type = doc_info.get('type', 'unknown')
        
        # Create categories and items from requirement results
        categories = {}
        
        for req_id, req_result in result.details.get('requirement_results', {}).items():
            req_data = req_result.get('requirement', {})
            category_id = req_data.get('category', 'Uncategorized')
            
            # Get or create category
            if category_id not in categories:
                categories[category_id] = {
                    'id': category_id,
                    'name': category_id,
                    'status': ValidationStatus.UNKNOWN.value,
                    'confidence_score': 0.0,
                    'items': [],
                    'errors': [],
                    'warnings': []
                }
            
            # Add item for this requirement
            compliance_level = req_result.get('compliance_level', ComplianceLevel.INDETERMINATE.value)
            if isinstance(compliance_level, ComplianceLevel):
                compliance_level = compliance_level.value
                
            status = status_map.get(
                ComplianceLevel(compliance_level), 
                ValidationStatus.UNKNOWN
            ).value
            
            item = {
                'id': req_id,
                'name': req_data.get('description', req_id),
                'status': status,
                'confidence_score': req_result.get('confidence_score', 0.5),
                'details': {
                    'requirement': req_data,
                    'justification': req_result.get('justification', ''),
                    'matched_keywords': req_result.get('matched_keywords', []),
                    'missing_keywords': req_result.get('missing_keywords', [])
                },
                'errors': [],
                'warnings': []
            }
            
            categories[category_id]['items'].append(item)
        
        # Determine category statuses based on items
        for cat_id, category in categories.items():
            statuses = [item['status'] for item in category['items']]
            if all(status == ValidationStatus.PASSED.value for status in statuses):
                category['status'] = ValidationStatus.PASSED.value
            elif any(status == ValidationStatus.FAILED.value for status in statuses):
                category['status'] = ValidationStatus.FAILED.value
            elif any(status == ValidationStatus.PARTIAL.value for status in statuses):
                category['status'] = ValidationStatus.PARTIAL.value
            else:
                category['status'] = ValidationStatus.UNKNOWN.value
                
            # Calculate average confidence
            if category['items']:
                category['confidence_score'] = sum(
                    item['confidence_score'] for item in category['items']
                ) / len(category['items'])
        
        # Create validation result structure
        overall_status = ValidationStatus.UNKNOWN
        if result.is_compliant:
            mode = result.details.get('mode_used', 'dynamic')
            
            # For dynamic mode, translate is_compliant to status
            if mode == 'dynamic':
                overall_status = ValidationStatus.PASSED if result.is_compliant else ValidationStatus.FAILED
            # For static mode, get more detailed status
            else:
                compliance_level = result.details.get('overall_compliance', ComplianceLevel.INDETERMINATE.value)
                if isinstance(compliance_level, ComplianceLevel):
                    compliance_level = compliance_level.value
                    
                status = status_map.get(
                    ComplianceLevel(compliance_level), 
                    ValidationStatus.UNKNOWN
                )
                overall_status = status
                
        # Build the final structure
        timestamp = result.details.get('timestamp', datetime.now().timestamp())
        
        formatted_result = {
            'document_id': doc_id,
            'document_name': doc_name,
            'document_type': doc_type,
            'status': overall_status.value,
            'metadata': {
                'timestamp': timestamp,
                'validator_version': '1.0.0',
                'mode': result.details.get('mode_used', 'dynamic'),
                'confidence_score': result.confidence,
                'processing_time_ms': result.details.get('processing_time', 0) * 1000,
                'warnings': []
            },
            'categories': list(categories.values()),
            'errors': [],
            'warnings': []
        }
        
        return formatted_result
    
    def _convert_validation_result(self, result: ValidationResult) -> DocumentComplianceReport:
        """Convert a ValidationResult to DocumentComplianceReport format"""
        # Map ValidationStatus to ComplianceLevel
        level_map = {
            ValidationStatus.PASSED: ComplianceLevel.FULLY_COMPLIANT,
            ValidationStatus.PARTIAL: ComplianceLevel.PARTIALLY_COMPLIANT,
            ValidationStatus.FAILED: ComplianceLevel.NON_COMPLIANT,
            ValidationStatus.UNKNOWN: ComplianceLevel.INDETERMINATE,
            ValidationStatus.ERROR: ComplianceLevel.INDETERMINATE
        }
        
        # Create requirements and requirement results
        requirements = []
        requirement_results = {}
        
        for category in result.categories:
            for item in category.items:
                # Create a requirement object
                req = {
                    'id': item.id,
                    'description': item.name,
                    'category': category.name,
                    'type': 'mandatory',  # Default type
                    'priority': 'medium',  # Default priority
                    'confidence_score': item.confidence_score
                }
                requirements.append(req)
                
                # Create requirement result
                compliance_level = level_map.get(item.status, ComplianceLevel.INDETERMINATE)
                req_result = {
                    'requirement': req,
                    'compliance_level': compliance_level.value,
                    'confidence_score': item.confidence_score,
                    'justification': item.details.get('justification', ''),
                    'matched_keywords': item.details.get('matched_keywords', []),
                    'missing_keywords': item.details.get('missing_keywords', [])
                }
                requirement_results[item.id] = req_result
        
        # Determine overall compliance
        if result.status == ValidationStatus.PASSED:
            overall_compliance = ComplianceLevel.FULLY_COMPLIANT
        elif result.status == ValidationStatus.PARTIAL:
            overall_compliance = ComplianceLevel.PARTIALLY_COMPLIANT
        elif result.status == ValidationStatus.FAILED:
            overall_compliance = ComplianceLevel.NON_COMPLIANT
        else:
            overall_compliance = ComplianceLevel.INDETERMINATE
            
        # Create the report
        report = {
            'document_id': result.document_id,
            'document_info': {
                'id': result.document_id,
                'name': result.document_name,
                'type': result.document_type
            },
            'requirements': requirements,
            'overall_compliance': overall_compliance.value,
            'compliance_confidence': result.metadata.confidence_score,
            'requirement_results': requirement_results,
            'timestamp': result.metadata.timestamp,
            'processing_time': result.metadata.processing_time_ms / 1000,  # Convert to seconds
            'mode_used': result.metadata.mode
        }
        
        return report
    
    def _map_output_format(self, format: OutputFormat) -> MatrixOutputFormat:
        """Map our OutputFormat to ComplianceMatrixGenerator's OutputFormat"""
        format_map = {
            OutputFormat.JSON: MatrixOutputFormat.JSON,
            OutputFormat.CSV: MatrixOutputFormat.CSV,
            OutputFormat.HTML: MatrixOutputFormat.HTML,
            OutputFormat.MARKDOWN: MatrixOutputFormat.MARKDOWN,
            OutputFormat.EXCEL: MatrixOutputFormat.EXCEL
        }
        return format_map.get(format, MatrixOutputFormat.JSON)
    
    def _convert_document_to_csv(self, result: Dict[str, Any]) -> str:
        """Convert a document result to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(["Document ID", result["document_id"]])
        writer.writerow(["Document Name", result["document_name"]])
        writer.writerow(["Document Type", result["document_type"]])
        writer.writerow(["Overall Status", result["status"]])
        writer.writerow(["Confidence", f"{result['metadata']['confidence_score']:.2f}"])
        writer.writerow(["Mode", result["metadata"]["mode"]])
        writer.writerow(["Timestamp", 
                        datetime.fromtimestamp(result["metadata"]["timestamp"]).isoformat()])
        writer.writerow([])  # Empty row
        
        # Write categories and items
        writer.writerow(["Category", "Status", "Confidence", "Item ID", "Item Name", 
                         "Item Status", "Item Confidence", "Details"])
        
        for category in result["categories"]:
            # Write first item in category with category info
            if category["items"]:
                first_item = category["items"][0]
                details = ""
                if self.include_justifications and "details" in first_item:
                    details = first_item["details"].get("justification", "")
                
                writer.writerow([
                    category["name"],
                    category["status"],
                    f"{category['confidence_score']:.2f}",
                    first_item["id"],
                    first_item["name"],
                    first_item["status"],
                    f"{first_item['confidence_score']:.2f}",
                    details
                ])
                
                # Write remaining items with empty category cells
                for item in category["items"][1:]:
                    details = ""
                    if self.include_justifications and "details" in item:
                        details = item["details"].get("justification", "")
                    
                    writer.writerow([
                        "", "", "", 
                        item["id"], 
                        item["name"], 
                        item["status"], 
                        f"{item['confidence_score']:.2f}",
                        details
                    ])
            else:
                # Category with no items
                writer.writerow([
                    category["name"],
                    category["status"],
                    f"{category['confidence_score']:.2f}",
                    "", "", "", "", ""
                ])
                
            writer.writerow([])  # Empty row between categories
        
        # Get CSV content
        output.seek(0)
        return output.read()
    
    def _convert_document_to_html(self, result: Dict[str, Any]) -> str:
        """Convert a document result to HTML format"""
        # Status to color mapping
        status_colors = {
            "passed": "#4CAF50",      # Green
            "partial": "#FFC107",     # Amber
            "failed": "#F44336",      # Red
            "unknown": "#9E9E9E",     # Grey
            "error": "#9C27B0"        # Purple
        }
        
        # Status to symbol mapping
        status_symbols = {
            "passed": "✓",
            "partial": "⚠",
            "failed": "✗",
            "unknown": "?",
            "error": "!"
        }
        
        html = []
        html.append("<!DOCTYPE html>")
        html.append("<html>")
        html.append("<head>")
        html.append("    <meta charset='UTF-8'>")
        html.append("    <meta name='viewport' content='width=device-width, initial-scale=1.0'>")
        html.append(f"    <title>Compliance Report - {result['document_name']}</title>")
        html.append("    <style>")
        html.append("        body { font-family: Arial, sans-serif; margin: 20px; line-height: 1.5; }")
        html.append("        h1, h2, h3 { color: #333; }")
        html.append("        .header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; }")
        html.append("        .header-title { flex-grow: 1; }")
        html.append("        .header-meta { background: #f5f5f5; padding: 10px; border-radius: 5px; font-size: 0.9em; }")
        html.append("        .document-info { margin-bottom: 30px; background: #f5f5f5; padding: 15px; border-radius: 5px; }")
        html.append("        .document-info table { width: 100%; border-collapse: collapse; }")
        html.append("        .document-info th { text-align: left; font-weight: normal; color: #666; width: 30%; }")
        html.append("        .status-badge { display: inline-block; padding: 5px 10px; border-radius: 3px; color: white; font-weight: bold; text-transform: uppercase; font-size: 0.8em; }")
        html.append("        .category { margin-bottom: 30px; border: 1px solid #ddd; border-radius: 5px; }")
        html.append("        .category-header { display: flex; justify-content: space-between; align-items: center; padding: 10px 15px; background: #f9f9f9; border-bottom: 1px solid #ddd; }")
        html.append("        .category-name { font-weight: bold; font-size: 1.1em; margin-right: 10px; }")
        html.append("        .category-status { display: flex; align-items: center; }")
        html.append("        .items-table { width: 100%; border-collapse: collapse; }")
        html.append("        .items-table th, .items-table td { text-align: left; padding: 10px; border-bottom: 1px solid #eee; }")
        html.append("        .items-table th { background: #f5f5f5; font-weight: 600; }")
        html.append("        .items-table tr:last-child td { border-bottom: none; }")
        html.append("        .items-table tr:hover { background: #f9f9f9; }")
        html.append("        .tooltip { position: relative; cursor: help; }")
        html.append("        .tooltip .tooltiptext { visibility: hidden; width: 200px; background-color: #333; color: #fff; text-align: center;")
        html.append("            border-radius: 6px; padding: 10px; position: absolute; z-index: 1; bottom: 125%; left: 50%; margin-left: -100px;")
        html.append("            opacity: 0; transition: opacity 0.3s; font-size: 0.9em; }")
        html.append("        .tooltip:hover .tooltiptext { visibility: visible; opacity: 0.9; }")
        
        # Add status-specific styles
        for status, color in status_colors.items():
            html.append(f"        .status-{status} {{ background-color: {color}; }}")
        
        html.append("    </style>")
        html.append("</head>")
        html.append("<body>")
        
        # Header
        html.append("    <div class='header'>")
        html.append("        <div class='header-title'>")
        html.append(f"            <h1>Compliance Report</h1>")
        html.append("        </div>")
        html.append("        <div class='header-meta'>")
        html.append(f"            Generated: {datetime.fromtimestamp(result['metadata']['timestamp']).strftime('%Y-%m-%d %H:%M:%S')}")
        html.append("        </div>")
        html.append("    </div>")
        
        # Document Information
        html.append("    <div class='document-info'>")
        html.append("        <h2>Document Information</h2>")
        html.append("        <table>")
        html.append("            <tr>")
        html.append("                <th>Document Name:</th>")
        html.append(f"                <td>{result['document_name']}</td>")
        html.append("            </tr>")
        html.append("            <tr>")
        html.append("                <th>Document ID:</th>")
        html.append(f"                <td>{result['document_id']}</td>")
        html.append("            </tr>")
        html.append("            <tr>")
        html.append("                <th>Document Type:</th>")
        html.append(f"                <td>{result['document_type']}</td>")
        html.append("            </tr>")
        html.append("            <tr>")
        html.append("                <th>Overall Status:</th>")
        status = result["status"]
        status_symbol = status_symbols.get(status, "?")
        html.append(f"                <td><span class='status-badge status-{status}'>{status_symbol} {status.upper()}</span></td>")
        html.append("            </tr>")
        
        if self.include_confidence:
            html.append("            <tr>")
            html.append("                <th>Confidence Score:</th>")
            confidence = result["metadata"]["confidence_score"]
            html.append(f"                <td>{confidence:.2f}</td>")
            html.append("            </tr>")
        
        html.append("            <tr>")
        html.append("                <th>Validation Mode:</th>")
        html.append(f"                <td>{result['metadata']['mode']}</td>")
        html.append("            </tr>")
        html.append("        </table>")
        html.append("    </div>")
        
        # Categories and Items
        html.append("    <h2>Validation Results</h2>")
        
        for category in result["categories"]:
            cat_status = category["status"]
            cat_symbol = status_symbols.get(cat_status, "?")
            
            html.append(f"    <div class='category'>")
            html.append(f"        <div class='category-header'>")
            html.append(f"            <div class='category-name'>{category['name']}</div>")
            html.append(f"            <div class='category-status'>")
            html.append(f"                <span class='status-badge status-{cat_status}'>{cat_symbol} {cat_status.upper()}</span>")
            
            if self.include_confidence:
                confidence = category["confidence_score"]
                html.append(f"                <span style='margin-left: 10px;'>(Confidence: {confidence:.2f})</span>")
                
            html.append(f"            </div>")
            html.append(f"        </div>")
            
            if category["items"]:
                html.append(f"        <table class='items-table'>")
                html.append(f"            <thead>")
                html.append(f"                <tr>")
                html.append(f"                    <th>Item</th>")
                html.append(f"                    <th>Status</th>")
                
                if self.include_confidence:
                    html.append(f"                    <th>Confidence</th>")
                    
                if self.include_justifications:
                    html.append(f"                    <th>Justification</th>")
                    
                html.append(f"                </tr>")
                html.append(f"            </thead>")
                html.append(f"            <tbody>")
                
                for item in category["items"]:
                    item_status = item["status"]
                    item_symbol = status_symbols.get(item_status, "?")
                    
                    html.append(f"                <tr>")
                    html.append(f"                    <td>{item['name']}")
                    
                    # Add tooltip with item ID if details are included
                    if self.include_details:
                        html.append(f" <span class='tooltip'>ℹ")
                        html.append(f"                        <span class='tooltiptext'>ID: {item['id']}</span>")
                        html.append(f"                    </span>")
                    
                    html.append(f"                    </td>")
                    html.append(f"                    <td><span class='status-badge status-{item_status}'>{item_symbol} {item_status.upper()}</span></td>")
                    
                    if self.include_confidence:
                        confidence = item["confidence_score"]
                        html.append(f"                    <td>{confidence:.2f}</td>")
                        
                    if self.include_justifications:
                        justification = ""
                        if "details" in item and "justification" in item["details"]:
                            justification = item["details"]["justification"]
                        html.append(f"                    <td>{justification}</td>")
                        
                    html.append(f"                </tr>")
                
                html.append(f"            </tbody>")
                html.append(f"        </table>")
            else:
                html.append(f"        <p style='padding: 15px; color: #666;'>No items in this category.</p>")
                
            html.append(f"    </div>")
        
        # Errors and Warnings
        if result["errors"] or result["warnings"]:
            html.append("    <h2>Issues</h2>")
            
            if result["errors"]:
                html.append("    <div style='margin-bottom: 15px;'>")
                html.append("        <h3>Errors</h3>")
                html.append("        <ul>")
                for error in result["errors"]:
                    html.append(f"            <li style='color: #F44336;'>{error}</li>")
                html.append("        </ul>")
                html.append("    </div>")
                
            if result["warnings"]:
                html.append("    <div>")
                html.append("        <h3>Warnings</h3>")
                html.append("        <ul>")
                for warning in result["warnings"]:
                    html.append(f"            <li style='color: #FFC107;'>{warning}</li>")
                html.append("        </ul>")
                html.append("    </div>")
        
        # Metadata footer
        if self.include_metadata:
            html.append("    <div style='margin-top: 30px; border-top: 1px solid #eee; padding-top: 15px; color: #999; font-size: 0.8em;'>")
            html.append("        <p>")
            html.append(f"            Validator Version: {result['metadata']['validator_version']}<br>")
            html.append(f"            Processing Time: {result['metadata']['processing_time_ms']:.2f} ms<br>")
            html.append(f"            Timestamp: {datetime.fromtimestamp(result['metadata']['timestamp']).strftime('%Y-%m-%d %H:%M:%S')}")
            html.append("        </p>")
            html.append("    </div>")
        
        html.append("</body>")
        html.append("</html>")
        
        return "\n".join(html)
    
    def _convert_document_to_markdown(self, result: Dict[str, Any]) -> str:
        """Convert a document result to Markdown format"""
        # Status to symbol mapping
        status_symbols = {
            "passed": "✅",
            "partial": "⚠️",
            "failed": "❌",
            "unknown": "❓",
            "error": "⛔"
        }
        
        md = []
        
        # Document header
        md.append(f"# Compliance Report: {result['document_name']}")
        md.append(f"Generated: {datetime.fromtimestamp(result['metadata']['timestamp']).strftime('%Y-%m-%d %H:%M:%S')}")
        md.append("")
        
        # Document information
        md.append("## Document Information")
        md.append("")
        md.append(f"**Document ID:** {result['document_id']}")
        md.append(f"**Document Type:** {result['document_type']}")
        
        status = result["status"]
        status_symbol = status_symbols.get(status, "❓")
        md.append(f"**Overall Status:** {status_symbol} {status.upper()}")
        
        if self.include_confidence:
            confidence = result["metadata"]["confidence_score"]
            md.append(f"**Confidence Score:** {confidence:.2f}")
            
        md.append(f"**Validation Mode:** {result['metadata']['mode']}")
        md.append("")
        
        # Categories and Items
        md.append("## Validation Results")
        md.append("")
        
        for category in result["categories"]:
            cat_status = category["status"]
            cat_symbol = status_symbols.get(cat_status, "❓")
            
            md.append(f"### {category['name']} {cat_symbol} {cat_status.upper()}")
            
            if self.include_confidence:
                confidence = category["confidence_score"]
                md.append(f"Confidence: {confidence:.2f}")
                
            md.append("")
            
            if category["items"]:
                # Create table header
                table_header = ["Item", "Status"]
                if self.include_confidence:
                    table_header.append("Confidence")
                if self.include_justifications:
                    table_header.append("Justification")
                    
                md.append("| " + " | ".join(table_header) + " |")
                md.append("| " + " | ".join(["---"] * len(table_header)) + " |")
                
                for item in category["items"]:
                    item_status = item["status"]
                    item_symbol = status_symbols.get(item_status, "❓")
                    
                    row = [
                        f"{item['name']} (ID: {item['id']})" if self.include_details else item['name'],
                        f"{item_symbol} {item_status.upper()}"
                    ]
                    
                    if self.include_confidence:
                        confidence = item["confidence_score"]
                        row.append(f"{confidence:.2f}")
                        
                    if self.include_justifications:
                        justification = ""
                        if "details" in item and "justification" in item["details"]:
                            justification = item["details"]["justification"]
                        row.append(justification)
                        
                    md.append("| " + " | ".join(row) + " |")
            else:
                md.append("*No items in this category.*")
                
            md.append("")
        
        # Errors and Warnings
        if result["errors"] or result["warnings"]:
            md.append("## Issues")
            md.append("")
            
            if result["errors"]:
                md.append("### Errors")
                md.append("")
                for error in result["errors"]:
                    md.append(f"- ❌ {error}")
                md.append("")
                
            if result["warnings"]:
                md.append("### Warnings")
                md.append("")
                for warning in result["warnings"]:
                    md.append(f"- ⚠️ {warning}")
                md.append("")
        
        # Metadata footer
        if self.include_metadata:
            md.append("---")
            md.append(f"*Validator Version: {result['metadata']['validator_version']}*")
            md.append(f"*Processing Time: {result['metadata']['processing_time_ms']:.2f} ms*")
            
        return "\n".join(md)
    
    def _convert_document_to_excel(self, result: Dict[str, Any], output_path: Path) -> Path:
        """Convert a document result to Excel format"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.logger.error("pandas and openpyxl are required for Excel export.")
            raise ImportError("pandas and openpyxl are required for Excel export. "
                             "Please install with: pip install pandas openpyxl")
        
        # Create workbook and sheets
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        details_sheet = wb.create_sheet(title="Details")
        
        # Define styles
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='left', vertical='center')
        
        status_fills = {
            "passed": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
            "partial": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
            "failed": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),
            "unknown": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
            "error": PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        }
        
        # Add summary sheet content
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 50
        
        # Document information
        summary_sheet.append(["Compliance Report", result["document_name"]])
        summary_sheet.append(["Generated", datetime.fromtimestamp(result["metadata"]["timestamp"]).strftime('%Y-%m-%d %H:%M:%S')])
        summary_sheet.append([])
        
        summary_sheet.append(["Document Information", ""])
        summary_sheet.cell(row=4, column=1).font = header_font
        
        summary_sheet.append(["Document ID", result["document_id"]])
        summary_sheet.append(["Document Type", result["document_type"]])
        
        status_row = summary_sheet.row_dimensions[7]
        summary_sheet.append(["Overall Status", result["status"].upper()])
        summary_sheet.cell(row=7, column=2).fill = status_fills.get(result["status"])
        
        if self.include_confidence:
            summary_sheet.append(["Confidence Score", f"{result['metadata']['confidence_score']:.2f}"])
            
        summary_sheet.append(["Validation Mode", result["metadata"]["mode"]])
        summary_sheet.append([])
        
        # Summary statistics
        row = 11
        summary_sheet.append(["Validation Summary", ""])
        summary_sheet.cell(row=row, column=1).font = header_font
        row += 1
        
        # Count statuses
        status_counts = {"passed": 0, "partial": 0, "failed": 0, "unknown": 0, "error": 0}
        for category in result["categories"]:
            for item in category["items"]:
                status_counts[item["status"]] += 1
                
        total_items = sum(status_counts.values())
        
        for status, count in status_counts.items():
            if count > 0:
                percentage = (count / total_items) * 100 if total_items > 0 else 0
                summary_sheet.append([f"{status.title()} Items", f"{count} ({percentage:.1f}%)"])
                summary_sheet.cell(row=row, column=1).fill = status_fills.get(status)
                row += 1
        
        # Add category summary
        row += 2
        summary_sheet.append(["Category Summary", ""])
        summary_sheet.cell(row=row, column=1).font = header_font
        row += 1
        
        summary_sheet.append(["Category", "Status", "Confidence", "Items"])
        for i in range(1, 5):
            summary_sheet.cell(row=row, column=i).font = header_font
            summary_sheet.cell(row=row, column=i).fill = header_fill
        row += 1
        
        for category in result["categories"]:
            summary_sheet.append([
                category["name"],
                category["status"].upper(),
                f"{category['confidence_score']:.2f}" if "confidence_score" in category else "",
                len(category["items"])
            ])
            summary_sheet.cell(row=row, column=2).fill = status_fills.get(category["status"])
            row += 1
        
        # Create details sheet
        details_sheet.column_dimensions['A'].width = 15
        details_sheet.column_dimensions['B'].width = 30
        details_sheet.column_dimensions['C'].width = 15
        details_sheet.column_dimensions['D'].width = 15
        details_sheet.column_dimensions['E'].width = 50
        
        details_row = 1
        details_sheet.append(["Category", "Item", "Status", "Confidence", "Justification"])
        for i in range(1, 6):
            details_sheet.cell(row=details_row, column=i).font = header_font
            details_sheet.cell(row=details_row, column=i).fill = header_fill
        details_row += 1
        
        for category in result["categories"]:
            for item in category["items"]:
                justification = ""
                if self.include_justifications and "details" in item and "justification" in item["details"]:
                    justification = item["details"]["justification"]
                    
                details_sheet.append([
                    category["name"],
                    item["name"],
                    item["status"].upper(),
                    f"{item['confidence_score']:.2f}" if "confidence_score" in item else "",
                    justification
                ])
                
                details_sheet.cell(row=details_row, column=3).fill = status_fills.get(item["status"])
                details_row += 1
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _generate_document_summary(
        self, 
        result: Dict[str, Any], 
        output_format: OutputFormat,
        output_path: Optional[Path]
    ) -> Union[Dict[str, Any], str, Path]:
        """Generate a summary report for a single document"""
        # Extract key information
        summary_data = {
            "document_id": result.get("document_id", "unknown"),
            "document_name": result.get("document_name", "unknown"),
            "document_type": result.get("document_type", "unknown"),
            "status": result.get("status", "unknown"),
            "confidence": result["metadata"].get("confidence_score", 0.0),
            "mode": result["metadata"].get("mode", "unknown"),
            "timestamp": result["metadata"].get("timestamp", 0.0),
            "category_summary": {},
            "status_counts": {
                "passed": 0,
                "partial": 0,
                "failed": 0,
                "unknown": 0,
                "error": 0
            }
        }
        
        # Calculate category summaries
        for category in result.get("categories", []):
            cat_id = category.get("id", "unknown")
            cat_name = category.get("name", cat_id)
            
            item_statuses = {}
            for item in category.get("items", []):
                status = item.get("status", "unknown")
                if status not in item_statuses:
                    item_statuses[status] = 0
                item_statuses[status] += 1
                
                # Update overall counts
                if status in summary_data["status_counts"]:
                    summary_data["status_counts"][status] += 1
            
            summary_data["category_summary"][cat_name] = {
                "status": category.get("status", "unknown"),
                "confidence": category.get("confidence_score", 0.0),
                "item_count": len(category.get("items", [])),
                "item_statuses": item_statuses
            }
        
        # Format based on the requested output format
        if output_format == OutputFormat.JSON:
            if output_path:
                self._save_json(summary_data, output_path)
            return summary_data
            
        elif output_format == OutputFormat.CSV:
            csv_content = self._summary_to_csv(summary_data)
            if output_path:
                self._save_text(csv_content, output_path)
            return csv_content
            
        elif output_format == OutputFormat.HTML:
            html_content = self._summary_to_html(summary_data)
            if output_path:
                self._save_text(html_content, output_path)
            return html_content
            
        elif output_format == OutputFormat.MARKDOWN:
            md_content = self._summary_to_markdown(summary_data)
            if output_path:
                self._save_text(md_content, output_path)
            return md_content
            
        elif output_format == OutputFormat.EXCEL:
            if not output_path:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
                    output_path = Path(temp.name)
            
            self._summary_to_excel(summary_data, output_path)
            return output_path
            
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    
    def _generate_matrix_summary(
        self, 
        data: Dict[str, Any], 
        output_format: OutputFormat,
        output_path: Optional[Path]
    ) -> Union[Dict[str, Any], str, Path]:
        """Generate a summary report for a compliance matrix"""
        # This delegates to the matrix generator's functionality
        # Normalize reports if needed
        if isinstance(data, dict) and all(isinstance(v, (ValidationResult, dict)) for v in data.values()):
            # Convert reports to common format if needed
            normalized_reports = {}
            for doc_id, report in data.items():
                if isinstance(report, dict):
                    # Assume it's already a DocumentComplianceReport dict
                    normalized_reports[doc_id] = report
                elif isinstance(report, ValidationResult):
                    # Convert ValidationResult to DocumentComplianceReport
                    normalized_reports[doc_id] = self._convert_validation_result(report)
                else:
                    raise TypeError(f"Unsupported report type for document {doc_id}: {type(report)}")
                    
            matrix_format = self._map_output_format(output_format)
            
            # Generate matrix and extract summary
            matrix = self.matrix_generator.generate_matrix(
                reports=normalized_reports,
                output_format=MatrixOutputFormat.JSON
            )
            
            summary = matrix.get("summary", {})
            
            # If output format is JSON, return just the summary part
            if output_format == OutputFormat.JSON:
                if output_path:
                    self._save_json(summary, output_path)
                return summary
                
            # For other formats, use the matrix generator but with a filter for summary data
            result = self.matrix_generator.generate_matrix(
                reports=normalized_reports,
                output_format=matrix_format,
                output_path=output_path
            )
            
            return result
        else:
            # If data is already a matrix or summary, use it directly
            matrix_format = self._map_output_format(output_format)
            
            if output_format == OutputFormat.JSON:
                if output_path:
                    self._save_json(data, output_path)
                return data
                
            # For other formats, need to convert the summary to those formats
            # This could be implemented based on desired output structure
            # For simplicity, return JSON for now
            if output_path:
                self._save_json(data, output_path)
            return data
    
    def _summary_to_csv(self, summary: Dict[str, Any]) -> str:
        """Convert summary to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header information
        writer.writerow(["Document Summary"])
        writer.writerow(["Document ID", summary["document_id"]])
        writer.writerow(["Document Name", summary["document_name"]])
        writer.writerow(["Document Type", summary["document_type"]])
        writer.writerow(["Overall Status", summary["status"]])
        writer.writerow(["Confidence", f"{summary['confidence']:.2f}"])
        writer.writerow(["Mode", summary["mode"]])
        writer.writerow(["Timestamp", 
                        datetime.fromtimestamp(summary["timestamp"]).isoformat()])
        writer.writerow([])
        
        # Write status counts
        writer.writerow(["Status Counts"])
        for status, count in summary["status_counts"].items():
            writer.writerow([status.title(), count])
        writer.writerow([])
        
        # Write category summary
        writer.writerow(["Category", "Status", "Confidence", "Items", "Passed", "Partial", "Failed", "Unknown", "Error"])
        
        for cat_name, cat_data in summary["category_summary"].items():
            statuses = cat_data["item_statuses"]
            writer.writerow([
                cat_name,
                cat_data["status"],
                f"{cat_data['confidence']:.2f}",
                cat_data["item_count"],
                statuses.get("passed", 0),
                statuses.get("partial", 0),
                statuses.get("failed", 0),
                statuses.get("unknown", 0),
                statuses.get("error", 0)
            ])
        
        output.seek(0)
        return output.read()
    
    def _summary_to_html(self, summary: Dict[str, Any]) -> str:
        """Convert summary to HTML format"""
        # Simplified version that creates a clean, basic HTML summary
        status_colors = {
            "passed": "#4CAF50",      # Green
            "partial": "#FFC107",     # Amber
            "failed": "#F44336",      # Red
            "unknown": "#9E9E9E",     # Grey
            "error": "#9C27B0"        # Purple
        }
        
        html = []
        html.append("<!DOCTYPE html>")
        html.append("<html>")
        html.append("<head>")
        html.append("    <meta charset='UTF-8'>")
        html.append("    <meta name='viewport' content='width=device-width, initial-scale=1.0'>")
        html.append(f"    <title>Compliance Summary - {summary['document_name']}</title>")
        html.append("    <style>")
        html.append("        body { font-family: Arial, sans-serif; margin: 20px; line-height: 1.5; }")
        html.append("        h1, h2, h3 { color: #333; }")
        html.append("        .section { margin-bottom: 30px; }")
        html.append("        .status-badge { display: inline-block; padding: 5px 10px; border-radius: 3px; color: white; font-weight: bold; text-transform: uppercase; font-size: 0.8em; }")
        html.append("        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }")
        html.append("        th, td { text-align: left; padding: 8px; border-bottom: 1px solid #eee; }")
        html.append("        th { background: #f5f5f5; font-weight: 600; }")
        html.append("        .chart-container { position: relative; height: 200px; margin-bottom: 20px; }")
        
        # Add status-specific styles
        for status, color in status_colors.items():
            html.append(f"        .status-{status} {{ background-color: {color}; }}")
            
        html.append("    </style>")
        html.append("</head>")
        html.append("<body>")
        
        # Header
        html.append(f"    <h1>Compliance Summary</h1>")
        html.append(f"    <p>Generated: {datetime.fromtimestamp(summary['timestamp']).strftime('%Y-%m-%d %H:%M:%S')}</p>")
        
        # Document Information
        html.append("    <div class='section'>")
        html.append("        <h2>Document Information</h2>")
        html.append("        <table>")
        html.append("            <tr><th>Document Name</th><td>" + summary['document_name'] + "</td></tr>")
        html.append("            <tr><th>Document ID</th><td>" + summary['document_id'] + "</td></tr>")
        html.append("            <tr><th>Document Type</th><td>" + summary['document_type'] + "</td></tr>")
        html.append("            <tr><th>Overall Status</th><td>")
        html.append(f"                <span class='status-badge status-{summary['status']}'>{summary['status'].upper()}</span>")
        html.append("            </td></tr>")
        html.append("            <tr><th>Confidence</th><td>" + f"{summary['confidence']:.2f}" + "</td></tr>")
        html.append("            <tr><th>Mode</th><td>" + summary['mode'] + "</td></tr>")
        html.append("        </table>")
        html.append("    </div>")
        
        # Status Counts
        html.append("    <div class='section'>")
        html.append("        <h2>Validation Results</h2>")
        
        # Create status count table
        html.append("        <table>")
        html.append("            <tr><th>Status</th><th>Count</th><th>Percentage</th></tr>")
        
        total_items = sum(summary["status_counts"].values())
        for status, count in summary["status_counts"].items():
            if total_items > 0:
                percentage = (count / total_items) * 100
            else:
                percentage = 0
                
            html.append("            <tr>")
            html.append(f"                <td><span class='status-badge status-{status}'>{status.upper()}</span></td>")
            html.append(f"                <td>{count}</td>")
            html.append(f"                <td>{percentage:.1f}%</td>")
            html.append("            </tr>")
            
        html.append("        </table>")
        
        # Add visualization chart
        html.append("        <div class='chart-container'>")
        html.append("            <div style='display: flex; height: 100%;'>")
        
        for status, count in summary["status_counts"].items():
            if total_items > 0:
                width = (count / total_items) * 100
            else:
                width = 0
                
            html.append(f"                <div class='status-{status}' style='width: {width}%; height: 100%; position: relative;'>")
            if width >= 5:  # Only show text if bar is wide enough
                html.append(f"                    <div style='position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); color: white;'>{count}</div>")
            html.append(f"                </div>")
            
        html.append("            </div>")
        html.append("        </div>")
        html.append("    </div>")
        
        # Category Summary
        html.append("    <div class='section'>")
        html.append("        <h2>Category Summary</h2>")
        html.append("        <table>")
        html.append("            <tr>")
        html.append("                <th>Category</th>")
        html.append("                <th>Status</th>")
        html.append("                <th>Confidence</th>")
        html.append("                <th>Items</th>")
        html.append("                <th>Passed</th>")
        html.append("                <th>Partial</th>")
        html.append("                <th>Failed</th>")
        html.append("                <th>Other</th>")
        html.append("            </tr>")
        
        for cat_name, cat_data in summary["category_summary"].items():
            statuses = cat_data["item_statuses"]
            
            html.append("            <tr>")
            html.append(f"                <td>{cat_name}</td>")
            html.append(f"                <td><span class='status-badge status-{cat_data['status']}'>{cat_data['status'].upper()}</span></td>")
            html.append(f"                <td>{cat_data['confidence']:.2f}</td>")
            html.append(f"                <td>{cat_data['item_count']}</td>")
            html.append(f"                <td>{statuses.get('passed', 0)}</td>")
            html.append(f"                <td>{statuses.get('partial', 0)}</td>")
            html.append(f"                <td>{statuses.get('failed', 0)}</td>")
            
            other_count = statuses.get('unknown', 0) + statuses.get('error', 0)
            html.append(f"                <td>{other_count}</td>")
            html.append("            </tr>")
            
        html.append("        </table>")
        html.append("    </div>")
        
        html.append("</body>")
        html.append("</html>")
        
        return "\n".join(html)
    
    def _summary_to_markdown(self, summary: Dict[str, Any]) -> str:
        """Convert summary to Markdown format"""
        # Status to symbol mapping
        status_symbols = {
            "passed": "✅",
            "partial": "⚠️",
            "failed": "❌",
            "unknown": "❓",
            "error": "⛔"
        }
        
        md = []
        
        # Header
        md.append(f"# Compliance Summary: {summary['document_name']}")
        md.append(f"Generated: {datetime.fromtimestamp(summary['timestamp']).strftime('%Y-%m-%d %H:%M:%S')}")
        md.append("")
        
        # Document information
        md.append("## Document Information")
        md.append("")
        md.append(f"**Document ID:** {summary['document_id']}")
        md.append(f"**Document Type:** {summary['document_type']}")
        
        status = summary["status"]
        status_symbol = status_symbols.get(status, "❓")
        md.append(f"**Overall Status:** {status_symbol} {status.upper()}")
        md.append(f"**Confidence:** {summary['confidence']:.2f}")
        md.append(f"**Mode:** {summary['mode']}")
        md.append("")
        
        # Status counts
        md.append("## Validation Results")
        md.append("")
        
        # Create status count table
        md.append("| Status | Count | Percentage |")
        md.append("| ------ | ----- | ---------- |")
        
        total_items = sum(summary["status_counts"].values())
        for status, count in summary["status_counts"].items():
            if total_items > 0:
                percentage = (count / total_items) * 100
            else:
                percentage = 0
                
            status_symbol = status_symbols.get(status, "❓")
            md.append(f"| {status_symbol} {status.upper()} | {count} | {percentage:.1f}% |")
            
        md.append("")
        
        # Category summary
        md.append("## Category Summary")
        md.append("")
        md.append("| Category | Status | Confidence | Items | Passed | Partial | Failed | Other |")
        md.append("| -------- | ------ | ---------- | ----- | ------ | ------- | ------ | ----- |")
        
        for cat_name, cat_data in summary["category_summary"].items():
            statuses = cat_data["item_statuses"]
            status = cat_data["status"]
            status_symbol = status_symbols.get(status, "❓")
            
            other_count = statuses.get('unknown', 0) + statuses.get('error', 0)
            
            md.append(f"| {cat_name} | {status_symbol} {status.upper()} | {cat_data['confidence']:.2f} | "
                     f"{cat_data['item_count']} | {statuses.get('passed', 0)} | "
                     f"{statuses.get('partial', 0)} | {statuses.get('failed', 0)} | {other_count} |")
            
        return "\n".join(md)
    
    def _summary_to_excel(self, summary: Dict[str, Any], output_path: Path) -> Path:
        """Convert summary to Excel format"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.charts import BarChart, Reference, PieChart
        except ImportError:
            self.logger.error("pandas, openpyxl are required for Excel export.")
            raise ImportError("pandas, openpyxl are required for Excel export. "
                             "Please install with: pip install pandas openpyxl")
                             
        # Create workbook
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_font = Font(bold=True)
        
        status_fills = {
            "passed": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
            "partial": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
            "failed": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),
            "unknown": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
            "error": PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        }
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 30
        
        # Document information
        summary_sheet.append(["Compliance Summary", summary["document_name"]])
        summary_sheet.cell(row=1, column=1).font = header_font
        
        summary_sheet.append(["Generated", datetime.fromtimestamp(summary["timestamp"]).strftime('%Y-%m-%d %H:%M:%S')])
        summary_sheet.append([])
        
        summary_sheet.append(["Document Information", ""])
        summary_sheet.cell(row=4, column=1).font = header_font
        
        summary_sheet.append(["Document ID", summary["document_id"]])
        summary_sheet.append(["Document Type", summary["document_type"]])
        summary_sheet.append(["Overall Status", summary["status"].upper()])
        summary_sheet.cell(row=7, column=2).fill = status_fills.get(summary["status"])
        summary_sheet.append(["Confidence", f"{summary['confidence']:.2f}"])
        summary_sheet.append(["Mode", summary["mode"]])
        summary_sheet.append([])
        
        # Status counts
        row = 11
        summary_sheet.append(["Validation Results", ""])
        summary_sheet.cell(row=row, column=1).font = header_font
        row += 1
        
        summary_sheet.append(["Status", "Count", "Percentage"])
        for i in range(1, 4):
            summary_sheet.cell(row=row, column=i).font = header_font
            summary_sheet.cell(row=row, column=i).fill = header_fill
        row += 1
        
        # Calculate percentages
        total_items = sum(summary["status_counts"].values())
        chart_data_start_row = row
        
        for status, count in summary["status_counts"].items():
            if total_items > 0:
                percentage = (count / total_items) * 100
            else:
                percentage = 0
                
            summary_sheet.append([status.upper(), count, f"{percentage:.1f}%"])
            summary_sheet.cell(row=row, column=1).fill = status_fills.get(status)
            row += 1
            
        chart_data_end_row = row - 1
        
        # Category summary
        row += 2
        summary_sheet.append(["Category Summary", ""])
        summary_sheet.cell(row=row, column=1).font = header_font
        row += 1
        
        # Add category table headers
        summary_sheet.append(["Category", "Status", "Confidence", "Items", 
                           "Passed", "Partial", "Failed", "Other"])
        for i in range(1, 9):
            summary_sheet.cell(row=row, column=i).font = header_font
            summary_sheet.cell(row=row, column=i).fill = header_fill
        row += 1
        
        # Add category data
        for cat_name, cat_data in summary["category_summary"].items():
            statuses = cat_data["item_statuses"]
            other_count = statuses.get('unknown', 0) + statuses.get('error', 0)
            
            summary_sheet.append([
                cat_name,
                cat_data["status"].upper(),
                f"{cat_data['confidence']:.2f}",
                cat_data["item_count"],
                statuses.get('passed', 0),
                statuses.get('partial', 0),
                statuses.get('failed', 0),
                other_count
            ])
            
            summary_sheet.cell(row=row, column=2).fill = status_fills.get(cat_data["status"])
            row += 1
            
        # Create pie chart for status distribution
        pie = PieChart()
        labels = Reference(summary_sheet, min_col=1, min_row=chart_data_start_row, max_row=chart_data_end_row)
        data = Reference(summary_sheet, min_col=2, min_row=chart_data_start_row, max_row=chart_data_end_row)
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = "Status Distribution"
        
        # Add the chart to the worksheet
        summary_sheet.add_chart(pie, "D4")
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _save_json(self, data: Dict[str, Any], output_path: Path) -> None:
        """Save JSON data to a file"""
        # Ensure directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"Saved compliance data to {output_path} (JSON format)")
    
    def _save_text(self, content: str, output_path: Path) -> None:
        """Save text content to a file"""
        # Ensure directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        # Determine format for logging
        format_name = "text"
        extension = output_path.suffix.lower()
        if extension == ".csv":
            format_name = "CSV"
        elif extension == ".html":
            format_name = "HTML"
        elif extension == ".md":
            format_name = "Markdown"
            
        self.logger.info(f"Saved compliance data to {output_path} ({format_name} format)")