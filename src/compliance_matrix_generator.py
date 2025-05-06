"""
Compliance Matrix Generator Module
Generates comprehensive compliance matrices in multiple output formats
with advanced filtering, sorting, and visualization capabilities.
"""

import logging
import json
import csv
import os
from typing import Dict, List, Optional, Union, Any, Callable, Set, Tuple
from pathlib import Path
from datetime import datetime
import re
from enum import Enum
import io
import base64
import tempfile

from .requirement_store import RequirementStore
from .policy_requirement_manager import PolicyRequirementManager
from .compliance_evaluator import (
    ComplianceEvaluator, 
    ComplianceLevel, 
    ComplianceResult, 
    DocumentComplianceReport
)


class OutputFormat(str, Enum):
    """Supported output formats for compliance matrices"""
    JSON = "json"
    CSV = "csv"
    HTML = "html"
    MARKDOWN = "markdown"
    EXCEL = "excel"  # Optional, requires pandas/openpyxl


class SortOrder(str, Enum):
    """Sorting order options"""
    ASCENDING = "ascending"
    DESCENDING = "descending"


class VisualizationStyle(str, Enum):
    """Visual styles for matrix representation"""
    TEXT = "text"           # Simple text indicators
    SYMBOL = "symbol"       # Use symbols like ✓, ⚠, ✗
    COLOR = "color"         # Use color-coding (for HTML/EXCEL)
    COLORBLIND = "colorblind"  # Color-blind friendly palette


class ComplianceMatrixGenerator:
    """
    Generates comprehensive compliance matrices in multiple formats,
    with filtering, sorting, and visualization capabilities.
    """
    
    # Symbols for different compliance levels
    COMPLIANCE_SYMBOLS = {
        ComplianceLevel.FULLY_COMPLIANT: "✓",
        ComplianceLevel.PARTIALLY_COMPLIANT: "⚠",
        ComplianceLevel.NON_COMPLIANT: "✗",
        ComplianceLevel.NOT_APPLICABLE: "-",
        ComplianceLevel.INDETERMINATE: "?"
    }
    
    # Color codes for HTML/CSS representation
    COMPLIANCE_COLORS = {
        ComplianceLevel.FULLY_COMPLIANT: "#4CAF50",     # Green
        ComplianceLevel.PARTIALLY_COMPLIANT: "#FFC107", # Amber
        ComplianceLevel.NON_COMPLIANT: "#F44336",       # Red
        ComplianceLevel.NOT_APPLICABLE: "#9E9E9E",      # Grey
        ComplianceLevel.INDETERMINATE: "#2196F3"        # Blue
    }
    
    # Color-blind friendly palette
    COLORBLIND_COLORS = {
        ComplianceLevel.FULLY_COMPLIANT: "#018571",     # Teal
        ComplianceLevel.PARTIALLY_COMPLIANT: "#80CDC1", # Light teal
        ComplianceLevel.NON_COMPLIANT: "#D01C8B",       # Magenta
        ComplianceLevel.NOT_APPLICABLE: "#DFC27D",      # Beige
        ComplianceLevel.INDETERMINATE: "#7570B3"        # Purple
    }
    
    def __init__(
        self,
        evaluator: Optional[ComplianceEvaluator] = None,
        visualization_style: VisualizationStyle = VisualizationStyle.SYMBOL,
        include_justifications: bool = True,
        include_confidence: bool = True,
        include_metadata: bool = True
    ):
        """
        Initialize the compliance matrix generator
        
        Args:
            evaluator: Optional ComplianceEvaluator for accessing compliance data
            visualization_style: How to visually represent compliance levels
            include_justifications: Whether to include justification text
            include_confidence: Whether to include confidence scores
            include_metadata: Whether to include metadata
        """
        self.logger = logging.getLogger(__name__)
        self.evaluator = evaluator or ComplianceEvaluator()
        self.visualization_style = visualization_style
        self.include_justifications = include_justifications
        self.include_confidence = include_confidence
        self.include_metadata = include_metadata
    
    def generate_matrix(
        self,
        reports: Dict[str, DocumentComplianceReport],
        output_format: OutputFormat = OutputFormat.JSON,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_order: SortOrder = SortOrder.ASCENDING,
        output_path: Optional[Path] = None
    ) -> Union[Dict, str, Path]:
        """
        Generate a compliance matrix from the provided compliance reports.
        
        Args:
            reports: Dictionary mapping document IDs to DocumentComplianceReport objects
            output_format: Desired output format
            filters: Optional filters to apply (e.g., {"compliance_level": "fully_compliant"})
            sort_by: Field to sort by (e.g., "document_id", "compliance_level", "category")
            sort_order: Sorting order (ascending or descending)
            output_path: Optional path to save the output
            
        Returns:
            Based on output_format:
              - JSON: Dictionary containing the matrix
              - CSV/HTML/MARKDOWN: String containing the formatted output
              - EXCEL: Path to the saved Excel file
        """
        # Generate the base matrix structure using the evaluator
        base_matrix = self.evaluator.generate_compliance_matrix(reports)
        
        # Apply filters if provided
        filtered_matrix = self._apply_filters(base_matrix, filters or {})
        
        # Apply sorting if requested
        sorted_matrix = self._apply_sorting(filtered_matrix, sort_by, sort_order)
        
        # Convert to the requested output format
        try:
            if output_format == OutputFormat.JSON:
                result = sorted_matrix
                if output_path:
                    self._save_to_json(result, output_path)
                return result
                
            elif output_format == OutputFormat.CSV:
                result = self._convert_to_csv(sorted_matrix)
                if output_path:
                    self._save_to_file(result, output_path)
                return result
                
            elif output_format == OutputFormat.HTML:
                result = self._convert_to_html(sorted_matrix)
                if output_path:
                    self._save_to_file(result, output_path)
                return result
                
            elif output_format == OutputFormat.MARKDOWN:
                result = self._convert_to_markdown(sorted_matrix)
                if output_path:
                    self._save_to_file(result, output_path)
                return result
                
            elif output_format == OutputFormat.EXCEL:
                if not output_path:
                    # Create a temporary file if no path provided
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                        output_path = Path(temp_file.name)
                
                result = self._convert_to_excel(sorted_matrix, output_path)
                return result
            
            else:
                raise ValueError(f"Unsupported output format: {output_format}")
        
        except Exception as e:
            self.logger.error(f"Error generating matrix in {output_format} format: {e}")
            raise
    
    def _apply_filters(self, matrix: Dict, filters: Dict[str, Any]) -> Dict:
        """
        Apply filters to the compliance matrix.
        
        Args:
            matrix: Base compliance matrix
            filters: Dictionary of filters to apply
            
        Returns:
            Filtered compliance matrix
        """
        # If no filters, return the original matrix
        if not filters:
            return matrix
        
        filtered_matrix = dict(matrix)  # Shallow copy
        
        # Filter documents
        if "document_id" in filters:
            doc_pattern = re.compile(filters["document_id"], re.IGNORECASE)
            filtered_matrix["documents"] = [
                doc for doc in matrix["documents"]
                if doc_pattern.search(doc["id"])
            ]
            
            # Also filter compliance_matrix entries
            doc_ids = {doc["id"] for doc in filtered_matrix["documents"]}
            filtered_matrix["compliance_matrix"] = [
                entry for entry in matrix["compliance_matrix"]
                if entry["document_id"] in doc_ids
            ]
        
        # Filter by document type
        if "document_type" in filters:
            type_pattern = re.compile(filters["document_type"], re.IGNORECASE)
            filtered_matrix["documents"] = [
                doc for doc in filtered_matrix["documents"]
                if "type" in doc and type_pattern.search(doc["type"])
            ]
            
            # Filter compliance_matrix based on document IDs
            doc_ids = {doc["id"] for doc in filtered_matrix["documents"]}
            filtered_matrix["compliance_matrix"] = [
                entry for entry in filtered_matrix["compliance_matrix"]
                if entry["document_id"] in doc_ids
            ]
        
        # Filter requirements
        if "requirement_id" in filters:
            req_pattern = re.compile(filters["requirement_id"], re.IGNORECASE)
            filtered_matrix["requirements"] = [
                req for req in matrix["requirements"]
                if req_pattern.search(req["id"])
            ]
            
            # Also filter results in each compliance_matrix entry
            req_ids = {req["id"] for req in filtered_matrix["requirements"]}
            for entry in filtered_matrix["compliance_matrix"]:
                entry["results"] = {
                    k: v for k, v in entry["results"].items()
                    if k in req_ids
                }
        
        # Filter by requirement category
        if "category" in filters:
            cat_pattern = re.compile(filters["category"], re.IGNORECASE)
            filtered_matrix["requirements"] = [
                req for req in filtered_matrix["requirements"]
                if "category" in req and cat_pattern.search(req["category"])
            ]
            
            # Filter results
            req_ids = {req["id"] for req in filtered_matrix["requirements"]}
            for entry in filtered_matrix["compliance_matrix"]:
                entry["results"] = {
                    k: v for k, v in entry["results"].items()
                    if k in req_ids
                }
        
        # Filter by compliance level
        if "compliance_level" in filters:
            level = filters["compliance_level"]
            # This is more complex as we need to filter at result level
            
            # First, collect document-requirement pairs that match the level
            matching_pairs = []
            for entry in matrix["compliance_matrix"]:
                doc_id = entry["document_id"]
                for req_id, result in entry["results"].items():
                    if result["compliance_level"] == level:
                        matching_pairs.append((doc_id, req_id))
            
            # Then filter based on matching pairs
            if matching_pairs:
                matching_doc_ids = {doc_id for doc_id, _ in matching_pairs}
                matching_req_ids = {req_id for _, req_id in matching_pairs}
                
                # Filter documents
                filtered_matrix["documents"] = [
                    doc for doc in filtered_matrix["documents"]
                    if doc["id"] in matching_doc_ids
                ]
                
                # Filter requirements
                filtered_matrix["requirements"] = [
                    req for req in filtered_matrix["requirements"]
                    if req["id"] in matching_req_ids
                ]
                
                # Filter compliance matrix entries
                filtered_matrix["compliance_matrix"] = [
                    entry for entry in filtered_matrix["compliance_matrix"]
                    if entry["document_id"] in matching_doc_ids
                ]
                
                # Filter results within entries
                for entry in filtered_matrix["compliance_matrix"]:
                    entry["results"] = {
                        k: v for k, v in entry["results"].items()
                        if k in matching_req_ids and (entry["document_id"], k) in matching_pairs
                    }
        
        # Filter by confidence score
        if "min_confidence" in filters:
            min_conf = float(filters["min_confidence"])
            
            # Collect document-requirement pairs that meet confidence threshold
            high_conf_pairs = []
            for entry in matrix["compliance_matrix"]:
                doc_id = entry["document_id"]
                for req_id, result in entry["results"].items():
                    if result.get("confidence_score", 0) >= min_conf:
                        high_conf_pairs.append((doc_id, req_id))
            
            # Filter based on confidence threshold
            if high_conf_pairs:
                # Keep only results above threshold
                for entry in filtered_matrix["compliance_matrix"]:
                    entry["results"] = {
                        k: v for k, v in entry["results"].items()
                        if v.get("confidence_score", 0) >= min_conf
                    }
        
        # Recalculate summary statistics if needed
        if filters:
            # Regenerate these from filtered data
            filtered_matrix["summary"] = self._generate_summary_stats(filtered_matrix)
            
            # Update metadata
            filtered_matrix["metadata"]["filtered"] = True
            filtered_matrix["metadata"]["filters_applied"] = filters
        
        return filtered_matrix
    
    def _apply_sorting(
        self, 
        matrix: Dict, 
        sort_by: Optional[str], 
        sort_order: SortOrder
    ) -> Dict:
        """
        Apply sorting to the compliance matrix.
        
        Args:
            matrix: Compliance matrix to sort
            sort_by: Field to sort by
            sort_order: Sorting order
            
        Returns:
            Sorted compliance matrix
        """
        # If no sorting requested, return the original matrix
        if not sort_by:
            return matrix
        
        sorted_matrix = dict(matrix)  # Shallow copy
        
        # Define key functions for different sort fields
        sort_key_funcs = {
            "document_id": lambda doc: doc["id"],
            "document_type": lambda doc: doc.get("type", ""),
            "document_name": lambda doc: doc.get("name", doc["id"]),
            "overall_compliance": lambda doc: doc.get("overall_compliance", "indeterminate"),
            "confidence_score": lambda doc: doc.get("confidence_score", 0),
            "requirement_id": lambda req: req["id"],
            "requirement_type": lambda req: req.get("type", ""),
            "requirement_priority": lambda req: req.get("priority", ""),
            "category": lambda req: req.get("category", "")
        }
        
        # Get the sort key function
        if sort_by in sort_key_funcs:
            key_func = sort_key_funcs[sort_by]
            reverse = (sort_order == SortOrder.DESCENDING)
            
            # Apply sorting based on the sort field type
            if sort_by.startswith("document"):
                sorted_matrix["documents"] = sorted(
                    matrix["documents"], 
                    key=key_func, 
                    reverse=reverse
                )
                
                # Keep compliance_matrix entries in same order as documents
                doc_order = {d["id"]: i for i, d in enumerate(sorted_matrix["documents"])}
                sorted_matrix["compliance_matrix"] = sorted(
                    matrix["compliance_matrix"],
                    key=lambda entry: doc_order.get(entry["document_id"], float('inf'))
                )
                
            elif sort_by.startswith("requirement") or sort_by == "category":
                sorted_matrix["requirements"] = sorted(
                    matrix["requirements"], 
                    key=key_func, 
                    reverse=reverse
                )
        
        return sorted_matrix
    
    def _generate_summary_stats(self, matrix: Dict) -> Dict:
        """
        Generate summary statistics for the matrix.
        
        Args:
            matrix: Compliance matrix
            
        Returns:
            Dictionary with summary statistics
        """
        # Initialize summary structure
        summary = {
            "overall_compliance": None,
            "compliance_by_requirement": {},
            "compliance_by_document": {},
            "compliance_by_category": {}
        }
        
        # Collect compliance levels for all results
        compliance_counts = {level.value: 0 for level in ComplianceLevel}
        total_results = 0
        
        # Collect requirements by ID for easy lookup
        req_map = {req["id"]: req for req in matrix.get("requirements", [])}
        
        # Initialize counters for requirements and categories
        for req in matrix.get("requirements", []):
            req_id = req["id"]
            summary["compliance_by_requirement"][req_id] = {
                level.value: 0 for level in ComplianceLevel
            }
            
            # Initialize category counters if not already done
            category = req.get("category", "Unknown")
            if category not in summary["compliance_by_category"]:
                summary["compliance_by_category"][category] = {
                    level.value: 0 for level in ComplianceLevel
                }
        
        # Process all matrix entries
        for entry in matrix.get("compliance_matrix", []):
            doc_id = entry["document_id"]
            
            # Initialize document compliance counter
            summary["compliance_by_document"][doc_id] = {
                level.value: 0 for level in ComplianceLevel
            }
            
            # Count results for this document
            for req_id, result in entry.get("results", {}).items():
                level = result.get("compliance_level", "indeterminate")
                
                # Count in overall totals
                compliance_counts[level] += 1
                total_results += 1
                
                # Count by requirement
                if req_id in summary["compliance_by_requirement"]:
                    summary["compliance_by_requirement"][req_id][level] += 1
                
                # Count by document
                summary["compliance_by_document"][doc_id][level] += 1
                
                # Count by category (if we can determine the category)
                if req_id in req_map:
                    category = req_map[req_id].get("category", "Unknown")
                    if category in summary["compliance_by_category"]:
                        summary["compliance_by_category"][category][level] += 1
        
        # Calculate percentages and determine overall compliance
        if total_results > 0:
            compliance_percentages = {
                level: (count / total_results) * 100 
                for level, count in compliance_counts.items()
            }
            
            # Determine overall compliance level
            fully_compliant_pct = compliance_percentages.get(ComplianceLevel.FULLY_COMPLIANT.value, 0)
            partially_compliant_pct = compliance_percentages.get(ComplianceLevel.PARTIALLY_COMPLIANT.value, 0)
            
            if fully_compliant_pct >= 80:
                overall_level = ComplianceLevel.FULLY_COMPLIANT.value
            elif fully_compliant_pct + partially_compliant_pct >= 60:
                overall_level = ComplianceLevel.PARTIALLY_COMPLIANT.value
            else:
                overall_level = ComplianceLevel.NON_COMPLIANT.value
            
            summary["overall_compliance"] = {
                "level": overall_level,
                "percentages": compliance_percentages,
                "counts": compliance_counts
            }
        else:
            # No results to evaluate
            summary["overall_compliance"] = {
                "level": ComplianceLevel.INDETERMINATE.value,
                "percentages": {level.value: 0 for level in ComplianceLevel},
                "counts": compliance_counts
            }
        
        return summary
    
    def _convert_to_csv(self, matrix: Dict) -> str:
        """
        Convert the compliance matrix to CSV format.
        
        Args:
            matrix: Compliance matrix
            
        Returns:
            String containing CSV data
        """
        # Use StringIO to build CSV data in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Get documents and requirements for easier access
        documents = matrix.get("documents", [])
        requirements = matrix.get("requirements", [])
        
        # Generate header row: Empty cell, then requirement IDs
        header = ["Document ID"]
        
        # Add document metadata columns if requested
        if self.include_metadata:
            header.extend(["Document Type", "Document Name"])
        
        # Add requirement IDs as column headers
        header.extend([req["id"] for req in requirements])
        
        # Write header row
        writer.writerow(header)
        
        # Create a lookup for requirement IDs
        req_lookup = {req["id"]: i for i, req in enumerate(requirements)}
        
        # Create a mapping from document IDs to their results
        doc_results = {}
        for entry in matrix.get("compliance_matrix", []):
            doc_results[entry["document_id"]] = entry.get("results", {})
        
        # Generate rows for each document
        for doc in documents:
            doc_id = doc["id"]
            row = [doc_id]
            
            # Add document metadata if requested
            if self.include_metadata:
                row.extend([doc.get("type", ""), doc.get("name", "")])
            
            # Get this document's results
            results = doc_results.get(doc_id, {})
            
            # Add a cell for each requirement
            for req in requirements:
                req_id = req["id"]
                result = results.get(req_id, {})
                
                # Get the compliance level (default to "indeterminate")
                level = result.get("compliance_level", "indeterminate")
                
                # Format the cell based on visualization style
                if self.visualization_style == VisualizationStyle.SYMBOL:
                    cell = self.COMPLIANCE_SYMBOLS.get(level, "?")
                elif self.visualization_style == VisualizationStyle.TEXT:
                    cell = level
                else:
                    # Default to level name for CSV
                    cell = level
                
                # Add confidence if requested
                if self.include_confidence and "confidence_score" in result:
                    confidence = result["confidence_score"]
                    cell = f"{cell} ({confidence:.2f})"
                
                row.append(cell)
            
            # Write the row
            writer.writerow(row)
        
        # Add summary information
        writer.writerow([])  # Empty row as separator
        
        # Add overall compliance info
        overall = matrix.get("summary", {}).get("overall_compliance", {})
        writer.writerow(["Overall Compliance", overall.get("level", "indeterminate")])
        
        # Add counts for each compliance level
        counts = overall.get("counts", {})
        for level in ComplianceLevel:
            count = counts.get(level.value, 0)
            writer.writerow([f"{level.value} Count", count])
        
        # Get the CSV data as a string
        output.seek(0)
        return output.read()
    
    def _convert_to_html(self, matrix: Dict) -> str:
        """
        Convert the compliance matrix to HTML format.
        
        Args:
            matrix: Compliance matrix
            
        Returns:
            String containing HTML content
        """
        documents = matrix.get("documents", [])
        requirements = matrix.get("requirements", [])
        
        # Start building HTML content
        html = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            "    <meta charset='UTF-8'>",
            "    <meta name='viewport' content='width=device-width, initial-scale=1.0'>",
            "    <title>Compliance Matrix Report</title>",
            "    <style>",
            "        body { font-family: Arial, sans-serif; margin: 20px; }",
            "        table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }",
            "        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "        th { background-color: #f2f2f2; position: sticky; top: 0; }",
            "        tr:nth-child(even) { background-color: #f9f9f9; }",
            "        tr:hover { background-color: #f2f2f2; }",
            "        .header-row th { text-align: center; font-weight: bold; }",
            "        .summary-section { margin-top: 30px; border-top: 2px solid #ddd; padding-top: 20px; }",
            "        .summary-table { width: auto; margin-bottom: 30px; }",
            "        .summary-table th { text-align: right; font-weight: normal; }",
            "        .summary-table td { font-weight: bold; }",
            "        .category-section { margin-top: 20px; }",
            "        .category-header { font-weight: bold; font-size: 1.1em; margin-bottom: 10px; }",
            "        .fully_compliant { background-color: #4CAF50; color: white; text-align: center; }",
            "        .partially_compliant { background-color: #FFC107; text-align: center; }",
            "        .non_compliant { background-color: #F44336; color: white; text-align: center; }",
            "        .not_applicable { background-color: #9E9E9E; color: white; text-align: center; }",
            "        .indeterminate { background-color: #2196F3; color: white; text-align: center; }",
            "        .tooltip { position: relative; display: inline-block; }",
            "        .tooltip .tooltiptext { visibility: hidden; width: 200px; background-color: #555;",
            "                              color: #fff; text-align: center; border-radius: 6px;",
            "                              padding: 5px; position: absolute; z-index: 1;",
            "                              bottom: 125%; left: 50%; margin-left: -100px;",
            "                              opacity: 0; transition: opacity 0.3s; }",
            "        .tooltip:hover .tooltiptext { visibility: visible; opacity: 1; }",
            "    </style>",
            "</head>",
            "<body>",
            "    <h1>Compliance Matrix Report</h1>",
            "    <p>Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "</p>"
        ]
        
        # Add the main matrix table
        html.append("    <h2>Compliance Matrix</h2>")
        html.append("    <table>")
        
        # Table header
        html.append("        <tr class='header-row'>")
        html.append("            <th rowspan='2'>Document ID</th>")
        
        # Add document metadata columns if requested
        if self.include_metadata:
            html.append("            <th rowspan='2'>Document Type</th>")
            html.append("            <th rowspan='2'>Document Name</th>")
        
        html.append(f"            <th colspan='{len(requirements)}'>Requirements</th>")
        html.append("        </tr>")
        
        # Second header row with requirement IDs
        html.append("        <tr>")
        for req in requirements:
            req_id = req["id"]
            description = req.get("description", "")
            req_type = req.get("type", "")
            priority = req.get("priority", "")
            
            # Create tooltip with requirement details
            tooltip = f"ID: {req_id}<br>Type: {req_type}<br>Priority: {priority}<br>Description: {description}"
            html.append(f"            <th><div class='tooltip'>{req_id}<span class='tooltiptext'>{tooltip}</span></div></th>")
        html.append("        </tr>")
        
        # Create a mapping from document IDs to their results
        doc_results = {}
        for entry in matrix.get("compliance_matrix", []):
            doc_results[entry["document_id"]] = entry.get("results", {})
        
        # Generate rows for each document
        for doc in documents:
            doc_id = doc["id"]
            doc_type = doc.get("type", "")
            doc_name = doc.get("name", doc_id)
            
            html.append("        <tr>")
            html.append(f"            <td>{doc_id}</td>")
            
            # Add document metadata if requested
            if self.include_metadata:
                html.append(f"            <td>{doc_type}</td>")
                html.append(f"            <td>{doc_name}</td>")
            
            # Get this document's results
            results = doc_results.get(doc_id, {})
            
            # Add a cell for each requirement
            for req in requirements:
                req_id = req["id"]
                result = results.get(req_id, {})
                
                # Get the compliance level (default to "indeterminate")
                level = result.get("compliance_level", "indeterminate")
                
                # Get justification if available and requested
                justification = result.get("justification", "No justification provided")
                confidence = result.get("confidence_score", 0.0)
                
                # Format the cell based on visualization style
                if self.visualization_style == VisualizationStyle.SYMBOL:
                    symbol = self.COMPLIANCE_SYMBOLS.get(level, "?")
                    cell_content = symbol
                elif self.visualization_style == VisualizationStyle.TEXT:
                    cell_content = level
                elif self.visualization_style == VisualizationStyle.COLORBLIND:
                    symbol = self.COMPLIANCE_SYMBOLS.get(level, "?")
                    cell_content = symbol
                else:  # Default to COLOR
                    symbol = self.COMPLIANCE_SYMBOLS.get(level, "?")
                    cell_content = symbol
                
                # Add confidence score if requested
                if self.include_confidence:
                    cell_content = f"{cell_content} ({confidence:.2f})"
                
                # Create a tooltip with justification if requested
                if self.include_justifications:
                    tooltip = f"<span class='tooltiptext'>{justification}</span>"
                    cell_content = f"<div class='tooltip'>{cell_content}{tooltip}</div>"
                
                # Add the cell with appropriate styling
                html.append(f"            <td class='{level}'>{cell_content}</td>")
            
            html.append("        </tr>")
        
        html.append("    </table>")
        
        # Add summary section
        html.append("    <div class='summary-section'>")
        html.append("        <h2>Summary</h2>")
        
        # Overall compliance
        overall = matrix.get("summary", {}).get("overall_compliance", {})
        overall_level = overall.get("level", "indeterminate")
        
        html.append("        <table class='summary-table'>")
        html.append("            <tr>")
        html.append("                <th>Overall Compliance:</th>")
        html.append(f"                <td class='{overall_level}'>{overall_level.replace('_', ' ').title()}</td>")
        html.append("            </tr>")
        
        # Add counts and percentages for each compliance level
        counts = overall.get("counts", {})
        percentages = overall.get("percentages", {})
        
        for level in ComplianceLevel:
            level_value = level.value
            count = counts.get(level_value, 0)
            percentage = percentages.get(level_value, 0)
            
            html.append("            <tr>")
            html.append(f"                <th>{level_value.replace('_', ' ').title()}:</th>")
            html.append(f"                <td>{count} ({percentage:.1f}%)</td>")
            html.append("            </tr>")
        
        html.append("        </table>")
        
        # Add category breakdown
        html.append("        <h3>Compliance by Category</h3>")
        
        categories = matrix.get("summary", {}).get("compliance_by_category", {})
        for category, counts in categories.items():
            html.append("        <div class='category-section'>")
            html.append(f"            <div class='category-header'>{category}</div>")
            html.append("            <table class='summary-table'>")
            
            # Add counts for each compliance level in this category
            for level in ComplianceLevel:
                level_value = level.value
                count = counts.get(level_value, 0)
                
                html.append("                <tr>")
                html.append(f"                    <th>{level_value.replace('_', ' ').title()}:</th>")
                html.append(f"                    <td class='{level_value}'>{count}</td>")
                html.append("                </tr>")
            
            html.append("            </table>")
            html.append("        </div>")
        
        html.append("    </div>")
        
        # Add metadata section if requested
        if self.include_metadata:
            html.append("    <div class='summary-section'>")
            html.append("        <h2>Metadata</h2>")
            html.append("        <table class='summary-table'>")
            
            metadata = matrix.get("metadata", {})
            for key, value in metadata.items():
                if key not in ("filters_applied",):  # Skip complex metadata
                    html.append("            <tr>")
                    html.append(f"                <th>{key.replace('_', ' ').title()}:</th>")
                    html.append(f"                <td>{value}</td>")
                    html.append("            </tr>")
            
            html.append("        </table>")
            html.append("    </div>")
        
        # Close HTML document
        html.append("</body>")
        html.append("</html>")
        
        return "\n".join(html)
    
    def _convert_to_markdown(self, matrix: Dict) -> str:
        """
        Convert the compliance matrix to Markdown format.
        
        Args:
            matrix: Compliance matrix
            
        Returns:
            String containing Markdown content
        """
        documents = matrix.get("documents", [])
        requirements = matrix.get("requirements", [])
        
        # Start building Markdown content
        md = [
            "# Compliance Matrix Report",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## Compliance Matrix"
        ]
        
        # Table header
        header = ["Document ID"]
        
        # Add document metadata columns if requested
        if self.include_metadata:
            header.extend(["Document Type", "Document Name"])
        
        # Add requirement IDs
        header.extend([f"{req['id']}" for req in requirements])
        
        # Add the header divider
        divider = ["-" * len(h) for h in header]
        
        # Add the table header
        md.append("| " + " | ".join(header) + " |")
        md.append("| " + " | ".join(divider) + " |")
        
        # Create a mapping from document IDs to their results
        doc_results = {}
        for entry in matrix.get("compliance_matrix", []):
            doc_results[entry["document_id"]] = entry.get("results", {})
        
        # Generate rows for each document
        for doc in documents:
            doc_id = doc["id"]
            doc_type = doc.get("type", "")
            doc_name = doc.get("name", doc_id)
            
            row = [doc_id]
            
            # Add document metadata if requested
            if self.include_metadata:
                row.extend([doc_type, doc_name])
            
            # Get this document's results
            results = doc_results.get(doc_id, {})
            
            # Add a cell for each requirement
            for req in requirements:
                req_id = req["id"]
                result = results.get(req_id, {})
                
                # Get the compliance level (default to "indeterminate")
                level = result.get("compliance_level", "indeterminate")
                
                # Format the cell based on visualization style
                if self.visualization_style == VisualizationStyle.SYMBOL:
                    cell = self.COMPLIANCE_SYMBOLS.get(level, "?")
                elif self.visualization_style == VisualizationStyle.TEXT:
                    cell = level.replace("_", " ").title()
                else:
                    # Default to symbols for Markdown
                    cell = self.COMPLIANCE_SYMBOLS.get(level, "?")
                
                # Add confidence if requested
                if self.include_confidence and "confidence_score" in result:
                    confidence = result["confidence_score"]
                    cell = f"{cell} ({confidence:.2f})"
                
                row.append(cell)
            
            # Add the row
            md.append("| " + " | ".join(row) + " |")
        
        # Add a section break
        md.append("")
        md.append("## Summary")
        md.append("")
        
        # Add overall compliance
        overall = matrix.get("summary", {}).get("overall_compliance", {})
        overall_level = overall.get("level", "indeterminate").replace("_", " ").title()
        md.append(f"**Overall Compliance:** {overall_level}")
        md.append("")
        
        # Add counts and percentages for each compliance level
        counts = overall.get("counts", {})
        percentages = overall.get("percentages", {})
        
        md.append("| Compliance Level | Count | Percentage |")
        md.append("| --- | --- | --- |")
        
        for level in ComplianceLevel:
            level_name = level.value.replace("_", " ").title()
            count = counts.get(level.value, 0)
            percentage = percentages.get(level.value, 0)
            
            md.append(f"| {level_name} | {count} | {percentage:.1f}% |")
        
        # Add category breakdown
        md.append("")
        md.append("## Compliance by Category")
        
        categories = matrix.get("summary", {}).get("compliance_by_category", {})
        for category, counts in categories.items():
            md.append("")
            md.append(f"### {category}")
            md.append("")
            
            md.append("| Compliance Level | Count |")
            md.append("| --- | --- |")
            
            for level in ComplianceLevel:
                level_name = level.value.replace("_", " ").title()
                count = counts.get(level.value, 0)
                
                md.append(f"| {level_name} | {count} |")
        
        # Add metadata if requested
        if self.include_metadata:
            md.append("")
            md.append("## Metadata")
            md.append("")
            
            metadata = matrix.get("metadata", {})
            md.append("| Key | Value |")
            md.append("| --- | --- |")
            
            for key, value in metadata.items():
                if key not in ("filters_applied",):  # Skip complex metadata
                    key_name = key.replace("_", " ").title()
                    md.append(f"| {key_name} | {value} |")
        
        return "\n".join(md)
    
    def _convert_to_excel(self, matrix: Dict, output_path: Path) -> Path:
        """
        Convert the compliance matrix to Excel format.
        
        Args:
            matrix: Compliance matrix
            output_path: Path to save the Excel file
            
        Returns:
            Path to the saved Excel file
        """
        try:
            # Attempt to import pandas
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.logger.error("pandas and openpyxl are required for Excel export.")
            raise ImportError("pandas and openpyxl are required for Excel export. "
                              "Please install with: pip install pandas openpyxl")
        
        # Create a new workbook
        wb = Workbook()
        
        # Create main matrix sheet
        matrix_sheet = wb.active
        matrix_sheet.title = "Compliance Matrix"
        
        # Get data for easier access
        documents = matrix.get("documents", [])
        requirements = matrix.get("requirements", [])
        
        # Define color fills for compliance levels
        if self.visualization_style == VisualizationStyle.COLORBLIND:
            fills = {
                "fully_compliant": PatternFill(start_color="018571", end_color="018571", fill_type="solid"),
                "partially_compliant": PatternFill(start_color="80CDC1", end_color="80CDC1", fill_type="solid"),
                "non_compliant": PatternFill(start_color="D01C8B", end_color="D01C8B", fill_type="solid"),
                "not_applicable": PatternFill(start_color="DFC27D", end_color="DFC27D", fill_type="solid"),
                "indeterminate": PatternFill(start_color="7570B3", end_color="7570B3", fill_type="solid")
            }
        else:
            fills = {
                "fully_compliant": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
                "partially_compliant": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
                "non_compliant": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),
                "not_applicable": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
                "indeterminate": PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            }
        
        # Define header styles
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_font = Font(bold=True)
        
        # Create headers
        headers = ["Document ID"]
        if self.include_metadata:
            headers.extend(["Document Type", "Document Name"])
        headers.extend([req["id"] for req in requirements])
        
        # Add headers to first row
        matrix_sheet.append(headers)
        
        # Style the headers
        for cell in matrix_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Create a mapping from document IDs to their results
        doc_results = {}
        for entry in matrix.get("compliance_matrix", []):
            doc_results[entry["document_id"]] = entry.get("results", {})
        
        # Add document rows
        for doc in documents:
            doc_id = doc["id"]
            doc_type = doc.get("type", "")
            doc_name = doc.get("name", doc_id)
            
            row = [doc_id]
            if self.include_metadata:
                row.extend([doc_type, doc_name])
            
            # Get this document's results
            results = doc_results.get(doc_id, {})
            
            # Add a cell for each requirement
            for req in requirements:
                req_id = req["id"]
                result = results.get(req_id, {})
                
                # Get the compliance level (default to "indeterminate")
                level = result.get("compliance_level", "indeterminate")
                
                # Format the cell based on visualization style
                if self.visualization_style == VisualizationStyle.SYMBOL:
                    cell_value = self.COMPLIANCE_SYMBOLS.get(level, "?")
                elif self.visualization_style == VisualizationStyle.TEXT:
                    cell_value = level
                else:
                    # Use symbols with color
                    cell_value = self.COMPLIANCE_SYMBOLS.get(level, "?")
                
                # Add confidence if requested
                if self.include_confidence and "confidence_score" in result:
                    confidence = result["confidence_score"]
                    cell_value = f"{cell_value} ({confidence:.2f})"
                
                row.append(cell_value)
            
            # Add row to sheet
            matrix_sheet.append(row)
        
        # Apply styling to data cells
        start_col = 1 + (2 if self.include_metadata else 0)  # Account for metadata columns
        for row_idx, doc in enumerate(documents, start=2):  # Start from row 2 (after header)
            doc_id = doc["id"]
            results = doc_results.get(doc_id, {})
            
            for col_idx, req in enumerate(requirements, start=start_col + 1):
                req_id = req["id"]
                result = results.get(req_id, {})
                level = result.get("compliance_level", "indeterminate")
                
                # Apply fill color
                cell = matrix_sheet.cell(row=row_idx, column=col_idx)
                cell.fill = fills.get(level, fills["indeterminate"])
                
                # Apply white font for better contrast on dark backgrounds
                if level in ("fully_compliant", "non_compliant", "not_applicable", "indeterminate"):
                    cell.font = Font(color="FFFFFF")
        
        # Add summary sheet
        summary_sheet = wb.create_sheet(title="Summary")
        
        # Add overall compliance
        overall = matrix.get("summary", {}).get("overall_compliance", {})
        overall_level = overall.get("level", "indeterminate")
        
        summary_sheet.append(["Overall Compliance", overall_level])
        summary_sheet.cell(row=1, column=2).fill = fills.get(overall_level, fills["indeterminate"])
        
        # Add counts and percentages
        summary_sheet.append([])
        summary_sheet.append(["Compliance Level", "Count", "Percentage"])
        
        counts = overall.get("counts", {})
        percentages = overall.get("percentages", {})
        
        row_idx = 4
        for level in ComplianceLevel:
            level_value = level.value
            count = counts.get(level_value, 0)
            percentage = percentages.get(level_value, 0)
            
            summary_sheet.append([level_value, count, f"{percentage:.1f}%"])
            summary_sheet.cell(row=row_idx, column=1).fill = fills.get(level_value, fills["indeterminate"])
            
            # Apply white font for better contrast on dark backgrounds
            if level_value in ("fully_compliant", "non_compliant", "not_applicable", "indeterminate"):
                summary_sheet.cell(row=row_idx, column=1).font = Font(color="FFFFFF")
            
            row_idx += 1
        
        # Add category breakdown
        summary_sheet.append([])
        summary_sheet.append(["Category Breakdown"])
        summary_sheet.cell(row=row_idx+1, column=1).font = header_font
        
        categories = matrix.get("summary", {}).get("compliance_by_category", {})
        row_idx += 2
        
        for category, counts in categories.items():
            summary_sheet.append([category])
            summary_sheet.cell(row=row_idx, column=1).font = header_font
            row_idx += 1
            
            summary_sheet.append(["Compliance Level", "Count"])
            row_idx += 1
            
            for level in ComplianceLevel:
                level_value = level.value
                count = counts.get(level_value, 0)
                
                summary_sheet.append([level_value, count])
                summary_sheet.cell(row=row_idx, column=1).fill = fills.get(level_value, fills["indeterminate"])
                
                # Apply white font for better contrast on dark backgrounds
                if level_value in ("fully_compliant", "non_compliant", "not_applicable", "indeterminate"):
                    summary_sheet.cell(row=row_idx, column=1).font = Font(color="FFFFFF")
                
                row_idx += 1
            
            # Add spacing between categories
            summary_sheet.append([])
            row_idx += 1
        
        # Save workbook to file
        wb.save(output_path)
        
        return output_path
    
    def _save_to_json(self, data: Dict, output_path: Path) -> None:
        """Save JSON data to a file"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"Saved compliance matrix to {output_path} (JSON format)")
    
    def _save_to_file(self, content: str, output_path: Path) -> None:
        """Save string content to a file"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        # Determine the format for logging
        format_name = "unknown"
        if output_path.suffix.lower() == ".csv":
            format_name = "CSV"
        elif output_path.suffix.lower() == ".html":
            format_name = "HTML"
        elif output_path.suffix.lower() == ".md":
            format_name = "Markdown"
        
        self.logger.info(f"Saved compliance matrix to {output_path} ({format_name} format)")


# Example usage
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s"
    )
    
    # Initialize evaluator and generator
    evaluator = ComplianceEvaluator()
    generator = ComplianceMatrixGenerator(evaluator=evaluator)
    
    # Example document
    document = {
        "filename": "example_audit.pdf",
        "type": "audit_rfi",
        "content": "This is an example audit document that discusses password policies. "
                   "All passwords must be at least 12 characters long. "
                   "User accounts are reviewed monthly."
    }
    
    # Evaluate document
    report = evaluator.evaluate_document(document)
    reports = {document["filename"]: report}
    
    # Generate matrix in different formats
    json_matrix = generator.generate_matrix(reports, output_format=OutputFormat.JSON)
    csv_content = generator.generate_matrix(reports, output_format=OutputFormat.CSV)
    html_content = generator.generate_matrix(reports, output_format=OutputFormat.HTML)
    
    # Save to files
    output_dir = Path("outputs")
    output_dir.mkdir(exist_ok=True)
    
    generator.generate_matrix(reports, output_format=OutputFormat.JSON, 
                             output_path=output_dir / "compliance_matrix.json")
    generator.generate_matrix(reports, output_format=OutputFormat.CSV,
                             output_path=output_dir / "compliance_matrix.csv")
    generator.generate_matrix(reports, output_format=OutputFormat.HTML,
                             output_path=output_dir / "compliance_matrix.html")